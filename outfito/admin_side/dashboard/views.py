from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.utils import timezone
from .services import (
    get_sales_report_data, 
    get_chart_data, 
    get_top_selling_products, 
    get_top_selling_categories
)
import json
import openpyxl
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.core.paginator import Paginator

def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None

@login_required(login_url='admin_login')
def admin_dashboard(request):
    chart_period = request.GET.get('period', 'yearly')
    year_str = request.GET.get('year')
    year = None
    if year_str and year_str.isdigit():
        year = int(year_str)
    
    report_data = get_sales_report_data()
    chart_data = get_chart_data(chart_period, year=year)
    top_products = get_top_selling_products(limit=5)
    top_categories = get_top_selling_categories(limit=5)
    
    orders_data = []
    for val in chart_data['data']:
        orders_data.append(0) 
    
    context = {
        'total_revenue': report_data['total_revenue'],
        'total_orders': report_data['total_orders'],
        'products_sold': report_data['products_sold'],
        'chart_labels': json.dumps(chart_data['labels']),
        'chart_data': json.dumps(chart_data['data']),
        'chart_orders': json.dumps(chart_data.get('orders', [])),
        'chart_period': chart_period,
        'top_products': top_products,
        'top_categories': top_categories,
        'recent_orders': report_data['orders'][:5],
        'selected_year': year or timezone.now().year,
    }
    return render(request, 'admin/dashboard.html', context)


@login_required(login_url='admin_login')
def sales_report(request):
    filter_type = request.GET.get('filter', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    today = timezone.now().date()
    
    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday()) # Monday
        end_date = today
    elif filter_type == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
        
    report_data = get_sales_report_data(start_date, end_date)
    
    orders_list = report_data['orders']
    coupons_list = report_data['coupon_usage_details']
    
    orders_paginator = Paginator(orders_list, 10)
    coupons_paginator = Paginator(coupons_list, 10)
    
    page_orders = request.GET.get('page_orders')
    page_coupons = request.GET.get('page_coupons')
    
    orders_page = orders_paginator.get_page(page_orders)
    coupons_page = coupons_paginator.get_page(page_coupons)

    def smart_page_range(page_obj):
        current = page_obj.number
        total   = page_obj.paginator.num_pages
        delta   = 2  # pages on each side of current
        pages   = set()
        pages.add(1)
        pages.add(total)
        for i in range(max(1, current - delta), min(total, current + delta) + 1):
            pages.add(i)
        result = []
        prev = None
        for p in sorted(pages):
            if prev is not None and p - prev > 1:
                result.append(None) 
            result.append(p)
            prev = p
        return result

    orders_page_range = smart_page_range(orders_page)

    period_choices = [
        ('', 'Custom'),
        ('today', 'Today'),
        ('week',  'This Week'),
        ('month', 'This Month'),
        ('year',  'This Year'),
    ]

    context = {
        'total_revenue': report_data['total_revenue'],
        'total_orders': report_data['total_orders'],
        'products_sold': report_data['products_sold'],
        'total_discount': report_data['total_discount'],
        'coupons_used': report_data['coupons_used'],
        'coupon_usage_details': coupons_page,
        'orders': orders_page,
        'orders_page_range': orders_page_range,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'filter': filter_type,
        'period_choices': period_choices,
    }
    return render(request, 'admin/sales_report.html', context)

@login_required(login_url='admin_login')
def download_sales_report_excel(request):
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    filter_type = request.GET.get('filter', '')
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    today = timezone.now().date()
    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif filter_type == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today

    report_data = get_sales_report_data(start_date, end_date)
    orders = report_data['orders']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sales Report'
    
    columns = ['Order ID', 'Date', 'Customer', 'Total Amount', 'Discount', 'Final Amount']
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        
    for order in orders:
        row_num += 1
        row = [
            order.order_number,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.user.username,
            float(order.subtotal + order.tax_amount + order.delivery_charge), # Total before discount
            float(order.discount_amount),
            float(order.total_amount)
        ]
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num).value = cell_value
            
    row_num += 2
    worksheet.cell(row=row_num, column=1).value = "Summary"
    worksheet.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = "Total Orders:"
    worksheet.cell(row=row_num, column=2).value = report_data['total_orders']
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = "Total Revenue:"
    worksheet.cell(row=row_num, column=2).value = float(report_data['total_revenue'])
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = "Total Discounts:"
    worksheet.cell(row=row_num, column=2).value = float(report_data['total_discount'])
    row_num += 1
    worksheet.cell(row=row_num, column=1).value = "Coupons Used:"
    worksheet.cell(row=row_num, column=2).value = report_data['coupons_used']

    if report_data['coupon_usage_details']:
        row_num += 3
        worksheet.cell(row=row_num, column=1).value = "Coupon Usage Details"
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)
        
        row_num += 1
        coupon_columns = ['Coupon Code', 'Type', 'Discount', 'Times Used', 'Usage Limit', 'Status']
        for col_num, column_title in enumerate(coupon_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            
        for coupon in report_data['coupon_usage_details']:
            row_num += 1
            limit = f"{coupon['times_used']}/{coupon['coupon__usage_limit']}" if coupon['coupon__usage_limit'] else "Unlimited"
            status = "Active" if coupon['coupon__is_active'] else "Inactive"
            row = [
                coupon['coupon__code'],
                coupon['coupon__discount_type'].title(),
                float(coupon['coupon__discount_value']),
                coupon['times_used'],
                limit,
                status
            ]
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value

    workbook.save(response)
    return response

@login_required(login_url='admin_login')
def download_sales_report_pdf(request):
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    filter_type = request.GET.get('filter', '')
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    today = timezone.now().date()
    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif filter_type == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif filter_type == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today

    report_data = get_sales_report_data(start_date, end_date)
    orders = report_data['orders']
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph("Sales Report", styles['Title']))
    elements.append(Spacer(1, 12))
    
    date_range = "All Time"
    if start_date and end_date:
        date_range = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    elif start_date:
        date_range = f"From {start_date.strftime('%B %d, %Y')}"
    elif end_date:
        date_range = f"Until {end_date.strftime('%B %d, %Y')}"
        
    elements.append(Paragraph(f"Period: {date_range}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    summary_data = [
        ["Total Orders", str(report_data['total_orders'])],
        ["Total Revenue", f"Rs. {report_data['total_revenue']:.2f}"],
        ["Total Discounts", f"Rs. {report_data['total_discount']:.2f}"],
        ["Products Sold", str(report_data['products_sold'])],
        ["Coupons Used", str(report_data['coupons_used'])]
    ]
    summary_table = Table(summary_data, hAlign='LEFT')
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))
    
    data = [['Order ID', 'Date', 'Customer', 'Amount', 'Discount', 'Final']]
    for order in orders:
        data.append([
            order.order_number,
            order.created_at.strftime('%Y-%m-%d'),
            order.user.username,
            f"{(order.subtotal + order.tax_amount + order.delivery_charge):.2f}",
            f"{order.discount_amount:.2f}",
            f"{order.total_amount:.2f}"
        ])
        
    order_table = Table(data, hAlign='LEFT', colWidths=[80, 80, 100, 70, 70, 70])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')), # Header color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(order_table)
    
    if report_data['coupon_usage_details']:
        elements.append(Spacer(1, 24))
        elements.append(Paragraph("Coupon Usage Details", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        coupon_data = [['Coupon Code', 'Type', 'Discount', 'Times Used', 'Usage Limit', 'Status']]
        for coupon in report_data['coupon_usage_details']:
            limit = f"{coupon['times_used']}/{coupon['coupon__usage_limit']}" if coupon['coupon__usage_limit'] else "Unlimited"
            status = "Active" if coupon['coupon__is_active'] else "Inactive"
            discount_val = f"{coupon['coupon__discount_value']:.2f}"
            if coupon['coupon__discount_type'] == 'percentage':
                discount_val = f"{coupon['coupon__discount_value']:.0f}%"
            else:
                discount_val = f"Rs. {discount_val}"
                
            coupon_data.append([
                coupon['coupon__code'],
                coupon['coupon__discount_type'].title(),
                discount_val,
                str(coupon['times_used']),
                limit,
                status
            ])
            
        coupon_table = Table(coupon_data, hAlign='LEFT')
        coupon_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#107c57')), # Brand green
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(coupon_table)

    doc.build(elements)
    
    return response
